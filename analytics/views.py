from django.shortcuts import render

# Create your views here.
from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse, HttpResponse
from django.db.models import Count, Q, F
from django.db.models.functions import TruncDate
from django.utils import timezone
from task.models import Task
from accounts.models import CustomUser
from .models import AnalyticsLog, DLQLog
import csv
import json
from datetime import datetime, timedelta
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment


@login_required
def analytics_dashboard(request):
    """Analytics dashboard"""
    if not request.user.is_admin():
        return JsonResponse({'error': 'Unauthorized'}, status=403)

    days = int(request.GET.get('days', 30))
    employee_filter = request.GET.get('employee', '')
    start_date = timezone.now() - timedelta(days=days)

    # Counts come from Task - the source of truth - not AnalyticsLog.
    # AnalyticsLog holds one row per lifecycle *event* (created/approved/
    # rejected), so counting its rows directly counts events, not tasks: a
    # task that's created then approved produced 2 AnalyticsLog rows and was
    # being counted as 2 "total tasks". AnalyticsLog is also populated by a
    # Kafka consumer, so it can lag or miss rows if that consumer was down.
    tasks = Task.objects.filter(created_at__gte=start_date)
    if employee_filter:
        tasks = tasks.filter(created_by__username=employee_filter)

    total_tasks = tasks.count()
    total_created = total_tasks  # every Task row is exactly one creation event
    total_approved = tasks.filter(status='approved').count()
    total_rejected = tasks.filter(status='rejected').count()

    success_rate = 0
    if total_approved + total_rejected > 0:
        success_rate = round((total_approved / (total_approved + total_rejected)) * 100, 2)

    employee_stats = tasks.annotate(employee_name=F('created_by__username')).values('employee_name').annotate(
        total=Count('id'),
        approved=Count('id', filter=Q(status='approved')),
        rejected=Count('id', filter=Q(status='rejected'))
    ).order_by('-total')

    daily_trends = tasks.filter(
        created_at__gte=timezone.now() - timedelta(days=7)
    ).annotate(
        date=TruncDate('created_at')
    ).values('date').annotate(
        total=Count('id'),
        approved=Count('id', filter=Q(status='approved')),
        rejected=Count('id', filter=Q(status='rejected'))
    ).order_by('date')

    # Recent Activity and the employee filter dropdown are audit/display data
    # (ip_address, user_agent, a per-event timeline), not counts, so the
    # event-log granularity from AnalyticsLog/CustomUser is exactly what's
    # wanted here - unlike the aggregates above, these aren't miscounted by
    # using event rows.
    recent_logs = AnalyticsLog.objects.filter(created_at__gte=start_date)
    if employee_filter:
        recent_logs = recent_logs.filter(employee_name=employee_filter)
    recent_logs = recent_logs[:50]
    all_employees = CustomUser.objects.filter(role='employee').values_list('username', flat=True)

    context = {
        'total_tasks': total_tasks,
        'total_created': total_created,
        'total_approved': total_approved,
        'total_rejected': total_rejected,
        'success_rate': success_rate,
        'employee_stats': employee_stats,
        'daily_trends': list(daily_trends),
        'recent_logs': recent_logs,
        'all_employees': all_employees,
        'selected_employee': employee_filter,
        'selected_days': days,
    }
    
    return render(request, 'analytics_dashboard.html', context)


# ⭐ NEW: DLQ Dashboard
@login_required
def dlq_dashboard(request):
    """
    Dashboard to monitor Dead Letter Queue
    """
    if not request.user.is_admin():
        return JsonResponse({'error': 'Unauthorized'}, status=403)
    
    # Get DLQ statistics
    total_dlq = DLQLog.objects.count()
    pending_dlq = DLQLog.objects.filter(status='pending').count()
    processing_dlq = DLQLog.objects.filter(status='processing').count()
    resolved_dlq = DLQLog.objects.filter(status='resolved').count()
    failed_dlq = DLQLog.objects.filter(status='failed').count()
    
    # Get recent DLQ entries
    recent_dlq = DLQLog.objects.all().order_by('-created_at')[:30]

    context = {
        'total_dlq': total_dlq,
        'pending_dlq': pending_dlq,
        'processing_dlq': processing_dlq,
        'resolved_dlq': resolved_dlq,
        'failed_dlq': failed_dlq,
        'recent_dlq': recent_dlq,
    }
    
    return render(request, 'dlq_dashboard.html', context)


# ⭐ NEW: Reprocess DLQ Entry
@login_required
def reprocess_dlq(request, dlq_id):
    """
    Manually reprocess a DLQ entry
    """
    if not request.user.is_admin():
        return JsonResponse({'error': 'Unauthorized'}, status=403)
    
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    
    try:
        dlq_entry = DLQLog.objects.get(id=dlq_id)
        DLQLog.objects.filter(id=dlq_id).update(status='processing')

        # Trigger reprocessing via Kafka (blocking - this is an explicit admin
        # action, so we want the real outcome, not a fire-and-forget submit).
        # Passing dlq_entry_id makes the producer update THIS row in place on
        # success/failure instead of creating a duplicate DLQ row - creating a
        # new row per failed retry was the bug that let one failed task
        # publish balloon into a dozen DLQ entries, each later replayed to
        # Kafka once it recovered, which is what inflated notification counts.
        from task.producer import producer
        success = producer.send_with_retry_sync(
            topic=dlq_entry.original_topic,
            message=dlq_entry.original_message,
            task_id=dlq_entry.task_id,
            dlq_entry_id=dlq_entry.id,
        )

        dlq_entry.refresh_from_db()
        if success:
            return JsonResponse({
                'success': True,
                'message': 'Message reprocessed successfully'
            })
        else:
            return JsonResponse({
                'success': False,
                'message': f'Reprocessing failed: {dlq_entry.error}'
            }, status=500)
            
    except DLQLog.DoesNotExist:
        return JsonResponse({'error': 'DLQ entry not found'}, status=404)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


# Download functions
@login_required
def download_csv(request):
    if not request.user.is_admin():
        return JsonResponse({'error': 'Unauthorized'}, status=403)
    
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="analytics_data.csv"'
    
    writer = csv.writer(response)
    writer.writerow(['Event Type', 'Task ID', 'Task Title', 'Employee', 'Status', 'Created At', 'IP Address'])
    
    logs = AnalyticsLog.objects.all().order_by('-created_at')
    for log in logs:
        writer.writerow([
            log.get_event_type_display(),
            log.task_id,
            log.task_title,
            log.employee_name,
            log.status,
            log.created_at.strftime('%Y-%m-%d %H:%M:%S'),
            log.ip_address or 'N/A'
        ])
    
    return response


@login_required
def download_excel(request):
    if not request.user.is_admin():
        return JsonResponse({'error': 'Unauthorized'}, status=403)
    
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "Analytics Data"
    
    headers = ['Event Type', 'Task ID', 'Task Title', 'Employee', 'Status', 'Created At', 'IP Address']
    for col, header in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    logs = AnalyticsLog.objects.all().order_by('-created_at')
    for row, log in enumerate(logs, 2):
        ws1.cell(row=row, column=1, value=log.get_event_type_display())
        ws1.cell(row=row, column=2, value=log.task_id)
        ws1.cell(row=row, column=3, value=log.task_title)
        ws1.cell(row=row, column=4, value=log.employee_name)
        ws1.cell(row=row, column=5, value=log.status)
        ws1.cell(row=row, column=6, value=log.created_at.strftime('%Y-%m-%d %H:%M:%S'))
        ws1.cell(row=row, column=7, value=log.ip_address or 'N/A')
    
    for column in ws1.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 30)
        ws1.column_dimensions[column_letter].width = adjusted_width
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="analytics_data.xlsx"'
    wb.save(response)
    return response


@login_required
def download_json(request):
    if not request.user.is_admin():
        return JsonResponse({'error': 'Unauthorized'}, status=403)
    
    logs = AnalyticsLog.objects.all().order_by('-created_at')
    data = []
    for log in logs:
        data.append({
            'event_type': log.get_event_type_display(),
            'task_id': log.task_id,
            'task_title': log.task_title,
            'employee': log.employee_name,
            'status': log.status,
            'created_at': log.created_at.strftime('%Y-%m-%d %H:%M:%S'),
            'ip_address': log.ip_address
        })
    
    response = HttpResponse(json.dumps(data, indent=2), content_type='application/json')
    response['Content-Disposition'] = 'attachment; filename="analytics_data.json"'
    return response